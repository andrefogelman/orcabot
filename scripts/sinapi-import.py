#!/usr/bin/env python3
"""
SINAPI Import Script
Parses official SINAPI Excel files from Caixa Economica Federal
and populates ob_sinapi_composicoes + ob_sinapi_composicao_insumos in Supabase.

Usage:
  python3 scripts/sinapi-import.py --file SINAPI_Preco_Ref_Insumos_SP_012026_Desonerado.xlsx --uf SP --data-base 2026-01 --tipo insumo --desoneracao com
  python3 scripts/sinapi-import.py --file SINAPI_Preco_Ref_Insumos_SP_012026_NaoDesonerado.xlsx --uf SP --data-base 2026-01 --tipo insumo --desoneracao sem
  python3 scripts/sinapi-import.py --file SINAPI_Custo_Ref_Composicoes_SP_012026_Desonerado.xlsx --uf SP --data-base 2026-01 --tipo composicao --desoneracao com

Requires: pip3 install openpyxl supabase python-dotenv
"""
import argparse
import os
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv()

SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    print('ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set in .env')
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


def parse_insumos(filepath: str, uf: str, data_base: str, desoneracao: str) -> list[dict]:
    """Parse SINAPI insumos (materials, labor, equipment) Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Find header row (contains 'CODIGO' or 'CODIGO')
    header_idx = None
    for i, row in enumerate(rows):
        row_str = ' '.join(str(c or '') for c in row).upper()
        if 'CODIGO' in row_str or 'CÓDIGO' in row_str:
            header_idx = i
            break

    if header_idx is None:
        print(f'ERROR: Could not find header row in {filepath}')
        return []

    headers = [str(c or '').strip().upper() for c in rows[header_idx]]

    # Map column names (SINAPI Excel varies slightly between months)
    col_map = {}
    for i, h in enumerate(headers):
        if 'CODIGO' in h or 'CÓDIGO' in h:
            col_map['codigo'] = i
        elif 'DESCRI' in h:
            col_map['descricao'] = i
        elif 'UNIDADE' in h or 'UN' == h:
            col_map['unidade'] = i
        elif 'PRECO' in h or 'PREÇO' in h or 'CUSTO' in h:
            if 'custo' not in col_map:
                col_map['custo'] = i

    if 'codigo' not in col_map or 'descricao' not in col_map:
        print(f'ERROR: Missing required columns. Found: {col_map}')
        return []

    records = []
    for row in rows[header_idx + 1:]:
        codigo_raw = row[col_map['codigo']] if col_map.get('codigo') is not None else None
        if not codigo_raw:
            continue
        codigo = str(codigo_raw).strip()
        if not codigo or not codigo[0].isdigit():
            continue

        descricao = str(row[col_map.get('descricao', 1)] or '').strip()
        if not descricao:
            continue

        unidade = str(row[col_map.get('unidade', 2)] or '').strip()
        custo_raw = row[col_map.get('custo', 3)] if col_map.get('custo') is not None else 0

        try:
            custo = float(str(custo_raw).replace('.', '').replace(',', '.')) if custo_raw else 0.0
        except (ValueError, TypeError):
            custo = 0.0

        # Classify by code range (SINAPI convention)
        classe = 'material'
        codigo_num = int(codigo) if codigo.isdigit() else 0
        if 25000 <= codigo_num < 44000:
            classe = 'mao_obra'
        elif 44000 <= codigo_num < 50000:
            classe = 'equipamento'

        record = {
            'codigo': codigo,
            'descricao': descricao,
            'unidade': unidade,
            'uf': uf,
            'data_base': data_base,
            'tipo': 'insumo',
            'classe': classe,
        }
        if desoneracao == 'com':
            record['custo_com_desoneracao'] = custo
            record['custo_sem_desoneracao'] = 0
        else:
            record['custo_com_desoneracao'] = 0
            record['custo_sem_desoneracao'] = custo

        records.append(record)

    wb.close()
    print(f'Parsed {len(records)} insumos from {filepath}')
    return records


def parse_composicoes(filepath: str, uf: str, data_base: str, desoneracao: str) -> list[dict]:
    """Parse SINAPI composicoes (service compositions) Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    header_idx = None
    for i, row in enumerate(rows):
        row_str = ' '.join(str(c or '') for c in row).upper()
        if 'CODIGO' in row_str or 'CÓDIGO' in row_str:
            header_idx = i
            break

    if header_idx is None:
        print(f'ERROR: Could not find header row in {filepath}')
        return []

    headers = [str(c or '').strip().upper() for c in rows[header_idx]]

    col_map = {}
    for i, h in enumerate(headers):
        if 'CODIGO' in h and 'COMPOSICAO' not in h:
            col_map['codigo'] = i
        elif 'COMPOSICAO' in h or 'COMPOSIÇÃO' in h:
            col_map['codigo_comp'] = i
        elif 'DESCRI' in h:
            col_map['descricao'] = i
        elif 'UNIDADE' in h or 'UN' == h:
            col_map['unidade'] = i
        elif 'CUSTO TOTAL' in h or 'PRECO TOTAL' in h or 'PREÇO TOTAL' in h:
            col_map['custo'] = i
        elif ('PRECO' in h or 'PREÇO' in h or 'CUSTO' in h) and 'custo' not in col_map:
            col_map['custo'] = i

    cod_col = col_map.get('codigo_comp', col_map.get('codigo'))
    if cod_col is None:
        print(f'ERROR: Missing codigo column. Found: {headers}')
        return []

    records = []
    for row in rows[header_idx + 1:]:
        codigo_raw = row[cod_col] if cod_col is not None else None
        if not codigo_raw:
            continue
        codigo = str(codigo_raw).strip()
        if not codigo or not codigo[0].isdigit():
            continue

        descricao = str(row[col_map.get('descricao', 1)] or '').strip()
        if not descricao:
            continue

        unidade = str(row[col_map.get('unidade', 2)] or '').strip()
        custo_raw = row[col_map.get('custo', 3)] if col_map.get('custo') is not None else 0

        try:
            custo = float(str(custo_raw).replace('.', '').replace(',', '.')) if custo_raw else 0.0
        except (ValueError, TypeError):
            custo = 0.0

        record = {
            'codigo': codigo,
            'descricao': descricao,
            'unidade': unidade,
            'uf': uf,
            'data_base': data_base,
            'tipo': 'composicao',
            'classe': 'material',  # compositions are mixed; default material
        }
        if desoneracao == 'com':
            record['custo_com_desoneracao'] = custo
            record['custo_sem_desoneracao'] = 0
        else:
            record['custo_com_desoneracao'] = 0
            record['custo_sem_desoneracao'] = custo

        records.append(record)

    wb.close()
    print(f'Parsed {len(records)} composicoes from {filepath}')
    return records


def upsert_records(records: list[dict], batch_size: int = 500) -> int:
    """Upsert records into ob_sinapi_composicoes table in batches."""
    total = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        result = supabase.table('ob_sinapi_composicoes').upsert(
            batch,
            on_conflict='codigo,uf,data_base',
        ).execute()
        total += len(batch)
        print(f'  Upserted batch {i // batch_size + 1}: {len(batch)} records (total: {total})')
    return total


def main():
    parser = argparse.ArgumentParser(description='Import SINAPI Excel into Supabase')
    parser.add_argument('--file', required=True, help='Path to SINAPI .xlsx file')
    parser.add_argument('--uf', required=True, help='State (e.g., SP, RJ, MG)')
    parser.add_argument('--data-base', required=True, help='Reference date (e.g., 2026-01)')
    parser.add_argument('--tipo', required=True, choices=['insumo', 'composicao'], help='Type of file')
    parser.add_argument('--desoneracao', required=True, choices=['com', 'sem'], help='With or without payroll tax relief')
    parser.add_argument('--dry-run', action='store_true', help='Parse only, do not write to database')
    args = parser.parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        print(f'ERROR: File not found: {filepath}')
        sys.exit(1)

    print(f'SINAPI Import: {args.tipo} | UF: {args.uf} | Data-base: {args.data_base} | Desoneracao: {args.desoneracao}')
    print(f'File: {filepath}')

    if args.tipo == 'insumo':
        records = parse_insumos(str(filepath), args.uf, args.data_base, args.desoneracao)
    else:
        records = parse_composicoes(str(filepath), args.uf, args.data_base, args.desoneracao)

    if not records:
        print('No records parsed. Check the file format.')
        sys.exit(1)

    print(f'Total records parsed: {len(records)}')

    if args.dry_run:
        print('DRY RUN: Skipping database write')
        # Print sample
        for r in records[:5]:
            print(f'  {r["codigo"]} | {r["descricao"][:60]} | {r["unidade"]} | R$ {r.get("custo_com_desoneracao", 0) or r.get("custo_sem_desoneracao", 0):.2f}')
        return

    total = upsert_records(records)
    print(f'Import complete: {total} records upserted to ob_sinapi_composicoes')


if __name__ == '__main__':
    main()
