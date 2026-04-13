#!/usr/bin/env python3
"""
Update ob_sinapi_composicoes.custo_sem_desoneracao from SINAPI Excel (CSD sheet).
Reads the official SINAPI_Referência spreadsheet and extracts SP costs without payroll tax relief.

Usage: python3 scripts/sinapi-update-sem-desoneracao.py
"""
import openpyxl
import re
import os
from supabase import create_client

# Load env
env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ[k.strip()] = v.strip().strip('"')

sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_KEY"])

XLSX = "/tmp/sinapi-check/SINAPI-2026-02-formato-xlsx/SINAPI_Referência_2026_02.xlsx"

print("Loading Excel...")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=False)

# --- CSD: Composições Sem Desoneração ---
ws = wb["CSD"]
print("Reading CSD sheet (Composições Sem Desoneração)...")

# SP column is index 54 (Custo R$)
SP_COL = 54
COD_COL = 1

updated = 0
skipped = 0
not_found = 0
batch = []

for row in ws.iter_rows(min_row=11, values_only=True):
    raw_code = row[COD_COL]
    sp_cost = row[SP_COL]

    if raw_code is None:
        continue

    # Extract numeric code from HYPERLINK formula or plain value
    code_str = str(raw_code)
    match = re.search(r'(\d{4,6})', code_str)
    if not match:
        continue

    codigo = match.group(1)

    # Parse cost
    try:
        cost = float(sp_cost) if sp_cost else 0.0
    except (ValueError, TypeError):
        cost = 0.0

    if cost <= 0:
        skipped += 1
        continue

    batch.append({"codigo": codigo, "custo": cost})

    if len(batch) >= 100:
        # Batch update
        for item in batch:
            resp = sb.from_("ob_sinapi_composicoes") \
                .update({"custo_sem_desoneracao": item["custo"]}) \
                .eq("codigo", item["codigo"]) \
                .eq("uf", "SP") \
                .eq("tipo", "composicao") \
                .execute()
            if resp.data:
                updated += len(resp.data)
            else:
                not_found += 1
        batch = []
        if updated % 500 == 0:
            print(f"  {updated} updated...")

# Flush remaining
for item in batch:
    resp = sb.from_("ob_sinapi_composicoes") \
        .update({"custo_sem_desoneracao": item["custo"]}) \
        .eq("codigo", item["codigo"]) \
        .eq("uf", "SP") \
        .eq("tipo", "composicao") \
        .execute()
    if resp.data:
        updated += len(resp.data)
    else:
        not_found += 1

print(f"\n--- CSD (Composições Sem Desoneração) ---")
print(f"Updated: {updated}")
print(f"Skipped (cost=0): {skipped}")
print(f"Not found in DB: {not_found}")

# --- ISD: Insumos Sem Desoneração ---
ws2 = wb["ISD"]
print("\nReading ISD sheet (Insumos Sem Desoneração)...")

# Find SP column in ISD (different layout - UFs are direct columns)
row4 = list(ws2.iter_rows(min_row=4, max_row=4, values_only=True))[0]
sp_col_isd = None
for i, h in enumerate(row4):
    if h == "SP":
        sp_col_isd = i
        break

if sp_col_isd is None:
    print("SP column not found in ISD sheet!")
else:
    print(f"SP column at index {sp_col_isd}")

    # ISD header row
    row6 = list(ws2.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    print(f"Headers: {row6[0]}, {row6[1]}, {row6[2]}, {row6[3]}")

    updated_ins = 0
    skipped_ins = 0

    for row in ws2.iter_rows(min_row=7, values_only=True):
        raw_code = row[0]  # CODIGO in col 0 for insumos
        if raw_code is None:
            continue

        code_str = str(raw_code)
        match = re.search(r'(\d{4,6})', code_str)
        if not match:
            continue

        codigo = match.group(1)

        try:
            cost = float(row[sp_col_isd]) if row[sp_col_isd] else 0.0
        except (ValueError, TypeError):
            cost = 0.0

        if cost <= 0:
            skipped_ins += 1
            continue

        resp = sb.from_("ob_sinapi_composicoes") \
            .update({"custo_sem_desoneracao": cost}) \
            .eq("codigo", codigo) \
            .eq("uf", "SP") \
            .eq("tipo", "insumo") \
            .execute()
        if resp.data:
            updated_ins += len(resp.data)

    print(f"\n--- ISD (Insumos Sem Desoneração) ---")
    print(f"Updated: {updated_ins}")
    print(f"Skipped (cost=0): {skipped_ins}")

print("\nDone!")
